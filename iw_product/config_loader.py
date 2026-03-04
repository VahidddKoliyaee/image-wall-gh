"""
Module 01: Config Loader
=========================
Reads IW-Product.xlsx and returns a config dictionary.
Replaces: Excel Import groups + Manual Inputs group.

Usage:
    from iw_product.config_loader import load_config
    config = load_config(r"C:\\path\\to\\IW-Product.xlsx")
"""

import os


def _safe_float(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _safe_int(val, default=1):
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def _safe_bool(val, default=False):
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().lower() in ("true", "yes", "1")
    return default


def _parse_material_thickness(material_str):
    """Extract thickness from '0.090\" AL ANOD' -> 0.090"""
    try:
        return float(material_str.split('"')[0].strip())
    except (ValueError, IndexError):
        return 0.0


def _read_iw_parameters(ws):
    """Read IWParameters sheet -> (params dict, dims dict)."""
    params = {}
    for row in range(1, ws.max_row + 1):
        key = ws.cell(row, 1).value
        val = ws.cell(row, 2).value
        if key:
            params[str(key).strip()] = val

    dims = {}
    for row in range(1, ws.max_row + 1):
        key_e = ws.cell(row, 5).value
        val_h = ws.cell(row, 8).value
        if key_e:
            dims[str(key_e).strip()] = val_h

    return params, dims


def _read_iw_product(ws):
    """Read IWProduct sheet (lookup tables)."""
    materials = []
    for row in range(3, ws.max_row + 1):
        val = ws.cell(row, 2).value
        if val:
            materials.append(str(val))
        else:
            break

    grid_patterns = []
    for row in range(3, ws.max_row + 1):
        val = ws.cell(row, 4).value
        if val:
            grid_patterns.append(str(val))
        else:
            break

    styles = []
    for row in range(3, ws.max_row + 1):
        val = ws.cell(row, 5).value
        if val:
            styles.append(str(val))
        else:
            break

    panel_sizes = []
    for row in range(3, ws.max_row + 1):
        w = ws.cell(row, 6).value
        l = ws.cell(row, 7).value
        if w and l:
            panel_sizes.append({"width": float(w), "length": float(l)})
        else:
            break

    surface_colors = {}
    for row in range(13, ws.max_row + 1):
        name = ws.cell(row, 15).value
        rgb = ws.cell(row, 16).value
        if name and rgb:
            surface_colors[str(name)] = str(rgb)

    return {
        "materials": materials,
        "grid_patterns": grid_patterns,
        "styles": styles,
        "panel_sizes": panel_sizes,
        "surface_colors": surface_colors,
    }


def _read_nonuniform_grid(ws):
    """Read IW-NonUniformGrid sheet."""
    if ws.max_row <= 1 and ws.max_column <= 1:
        return None
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        rows.append(list(row))
    return rows if rows else None


def load_config(excel_path):
    """
    Load full project configuration from IW-Product.xlsx.

    Args:
        excel_path: Path to the Excel file.

    Returns:
        dict with all project parameters.

    Raises:
        FileNotFoundError: If excel_path doesn't exist.
    """
    if not excel_path or not os.path.exists(excel_path):
        raise FileNotFoundError("Excel file not found: {}".format(excel_path))

    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    config = {}

    # ── IWParameters ──────────────────────────────────────────────
    if "IWParameters" in wb.sheetnames:
        params, dims = _read_iw_parameters(wb["IWParameters"])
        ws = wb["IWParameters"]
        material = str(params.get("Material", ""))
        style = str(params.get("Style", ""))

        config.update({
            "product_number":       str(params.get("Product Number", "")),
            "scope_name":           str(params.get("Scope Name", "")),
            "image_name":           str(params.get("Image Name", "")),
            "scope_bot_left":       str(params.get("Scope Bot Left (X,Y)", "0,0")),
            "material":             material,
            "material_thickness":   _parse_material_thickness(material),
            "surface":              str(params.get("Surface", "")),
            "style":                style,
            "product_style":        str(params.get("Product Style", "")),
            "panel_row_height":     _safe_float(params.get("Panel Row Height (in)")),
            "panel_col_width":      _safe_float(params.get("Panel Column Width (in)")),
            "qty_columns":          _safe_int(params.get("QTY of Columns")),
            "qty_rows":             _safe_int(params.get("QTY of Rows")),
            "invert_image":         _safe_bool(params.get("Invert Image")),
            "grid_pattern":         str(params.get("Grid Pattern", "Default")),
            "cross_seam":           str(params.get("Cross Seam", "Default")),
            "no_perf_level":        _safe_bool(params.get("No Perforation Level")),
            "smallest_perf":        _safe_float(params.get("Smallest Perforation"), 0.1875),
            "perf_spacing":         _safe_float(params.get("Perf Spacing"), 1.0),
            "min_bridge":           _safe_float(params.get("Minimum Bridge"), 0.125),
            "line_spacing_target":  _safe_float(params.get("Line Spacing Target"), 1.25),
            "min_bridge_along":     _safe_float(params.get("Min Bridge Along Line"), 0.25),
            "min_bridge_perp":      _safe_float(params.get("Min Bridge Perp to Line"), 0.25),
            "min_rectangle":        _safe_float(params.get("Min Rectangle"), 0.1875),
            "max_rectangle":        _safe_float(params.get("Max Rectangle"), 2.0),
            "driver_curve_1_pts":   str(params.get("Driver Curve 1 Pts", "")),
            "driver_curve_2_pts":   str(params.get("Driver Curve 2 Pts", "")),
            "mullion_size":         str(params.get("Mullion Size (if exist)", "")),
            "surface_rgb":          str(params.get("Surface RGB", "200,200,200")),
            "transparency":         _safe_bool(params.get("Transparency")),
            "vertical":             "[Vertical]" in style,
            "overall_height":       _safe_float(dims.get("Overall Scope Height")),
            "overall_width":        _safe_float(dims.get("Overall Scope Width")),
            "max_panel_length":     _safe_float(ws.cell(10, 6).value, 120.0),
            "max_panel_width":      _safe_float(ws.cell(13, 6).value, 40.0),
        })

    # ── IWProduct (lookups) ───────────────────────────────────────
    if "IWProduct" in wb.sheetnames:
        lookups = _read_iw_product(wb["IWProduct"])
        config["material_list"] = lookups["materials"]
        config["style_list"] = lookups["styles"]
        config["grid_pattern_list"] = lookups["grid_patterns"]
        config["panel_sizes"] = lookups["panel_sizes"]
        config["surface_colors"] = lookups["surface_colors"]

    # ── IW-NonUniformGrid ─────────────────────────────────────────
    if "IW-NonUniformGrid" in wb.sheetnames:
        config["nonuniform_grid"] = _read_nonuniform_grid(wb["IW-NonUniformGrid"])
    else:
        config["nonuniform_grid"] = None

    wb.close()
    return config


def parse_rgb(rgb_string):
    """Parse '201,205,197' -> (201, 205, 197)."""
    parts = str(rgb_string).split(",")
    return (int(parts[0].strip()), int(parts[1].strip()), int(parts[2].strip()))


def parse_point_list(pts_string):
    """Parse '[-38.963,-14.963]#[-14.963,10.581]' -> [(-38.963,-14.963), ...]."""
    if not pts_string:
        return []
    points = []
    for pt_str in str(pts_string).split("#"):
        pt_str = pt_str.strip().strip("[]")
        if "," in pt_str:
            parts = pt_str.split(",")
            points.append((float(parts[0].strip()), float(parts[1].strip())))
    return points


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "IW-Product.xlsx"
    config = load_config(path)
    print("Product: {} - {}".format(config["product_number"], config["scope_name"]))
    print("Material: {} (t={})".format(config["material"], config["material_thickness"]))
    print("Style: {}".format(config["style"]))
    print("Grid: {}x{} panels, {}x{} in".format(
        config["qty_columns"], config["qty_rows"],
        config["panel_col_width"], config["panel_row_height"]))
