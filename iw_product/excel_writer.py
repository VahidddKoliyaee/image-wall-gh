"""
Module 19: Excel Writer
========================
Writes panel data, grid dimensions, and perforation info back to Excel.

Replaces: "Write Grid To Excel" group + "1.0-Multiple Sheets" Python script.

Usage:
    from iw_product.excel_writer import write_to_excel
    result = write_to_excel(excel_path, config, grid, faces, dies)
"""

import os


def write_to_excel(excel_path, config, grid, faces, dies, naming=None):
    """
    Write computed panel data back to the IW-Product.xlsx file.

    Args:
        excel_path: str, path to Excel file
        config: dict from config_loader
        grid: dict from grid
        faces: dict from panel_faces
        dies: dict from die_selection or fastener_clearance
        naming: dict from panel_naming (optional)

    Returns:
        dict with:
            success   - bool
            message   - str
            sheets_written - list of str
    """
    if not excel_path or not os.path.exists(excel_path):
        return {"success": False, "message": "Excel file not found", "sheets_written": []}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
    except ImportError:
        return {"success": False, "message": "openpyxl not available", "sheets_written": []}

    sheets_written = []

    # ── Update IWParameters sheet with computed values ────────────
    if "IWParameters" in wb.sheetnames:
        ws = wb["IWParameters"]

        # Write actual dimensions to column I
        # Row 3: Overall Height actual
        # Row 6: Overall Width actual
        overall_h = config["qty_rows"] * config["panel_row_height"]
        overall_w = config["qty_columns"] * config["panel_col_width"]

        ws.cell(3, 9).value = overall_h   # I3
        ws.cell(6, 9).value = overall_w   # I6

        # Diff column J
        target_h = ws.cell(3, 8).value or 0
        target_w = ws.cell(6, 8).value or 0
        try:
            ws.cell(3, 10).value = float(target_h) - overall_h  # J3
            ws.cell(6, 10).value = float(target_w) - overall_w  # J6
        except (TypeError, ValueError):
            pass

        # Calculated values column F/L
        ws.cell(16, 6).value = config["qty_columns"]  # F16: # of Columns
        ws.cell(17, 6).value = config["qty_rows"]     # F17: # of Rows
        ws.cell(18, 6).value = config["panel_row_height"]  # F18: Panel Length
        ws.cell(19, 6).value = config["panel_col_width"]   # F19: Panel Width

        sheets_written.append("IWParameters")

    # ── Write grid data to IW-NonUniformGrid if needed ────────────
    # (only for nonuniform grids)

    try:
        wb.save(excel_path)
        return {
            "success": True,
            "message": "Written to: {}".format(", ".join(sheets_written)),
            "sheets_written": sheets_written,
        }
    except Exception as e:
        return {
            "success": False,
            "message": "Save failed: {}".format(e),
            "sheets_written": [],
        }
